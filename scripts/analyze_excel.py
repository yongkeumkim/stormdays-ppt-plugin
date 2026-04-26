"""
analyze_excel.py — Excel 파일을 파싱해 슬라이드 콘텐츠 JSON 생성

Usage: python analyze_excel.py <excel_path> [output_json_path]

Excel 시트 구조:
  config  : 프레젠테이션 기본 설정 (title, subtitle, date, company)
  slides  : 슬라이드 목록 (order, type, section_num, title)
  stat    : 큰 숫자 강조 슬라이드 데이터
  two_col : 두 컬럼 텍스트 슬라이드 데이터
  steps   : Step 카드 슬라이드 데이터
  metrics : 막대/진행률 성과 슬라이드 데이터
  chart   : 파이 차트 슬라이드 데이터
  table   : 표 슬라이드 데이터
  closing : 결론 슬라이드 데이터
"""

import sys
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl


def sheet_to_records(ws):
    """워크시트를 딕셔너리 리스트로 변환 (첫 행 = 헤더)"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        rec = {}
        for h, v in zip(headers, row):
            rec[h] = v if v is not None else ""
        records.append(rec)
    return records


def _parse_table_sheet(ws) -> list:
    """
    table 시트 파싱: section_num이 있는 행 = 슬라이드 정의 + 열 헤더
    이후 section_num 없는 행들 = 데이터 행 (다음 섹션 행 전까지)
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    col_headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

    tables = []
    current = None

    for row in rows[1:]:
        if all(v is None for v in row):
            continue

        rec = dict(zip(col_headers, [v if v is not None else "" for v in row]))
        sec_num = str(rec.get("section_num", "")).strip()

        if sec_num and sec_num != "":
            # 새 테이블 슬라이드 시작
            if current is not None:
                tables.append(current)
            # 열 헤더는 col1, col2, col3... 값에서 가져옴
            data_cols = [k for k in col_headers if k not in ("section_num", "title")]
            header_row = [str(rec.get(c, "")) for c in data_cols]
            current = {
                "section_num": sec_num,
                "title": str(rec.get("title", "")),
                "headers": header_row,
                "rows": [],
            }
        else:
            # 데이터 행 (현재 테이블에 추가)
            if current is not None:
                data_cols = [k for k in col_headers if k not in ("section_num", "title")]
                data_row = [str(rec.get(c, "")) for c in data_cols]
                if any(v for v in data_row):  # 빈 행 무시
                    current["rows"].append(data_row)

    if current is not None:
        tables.append(current)

    return tables


def parse_excel(excel_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_names = [s.lower() for s in wb.sheetnames]
    sheets = {s.lower(): wb[name] for s, name in zip(sheet_names, wb.sheetnames)}

    result = {}

    # ── config ────────────────────────────────────────────────────────────────
    if "config" in sheets:
        rows = list(sheets["config"].iter_rows(values_only=True))
        config = {}
        for row in rows:
            if row[0] and row[1]:
                config[str(row[0]).strip().lower()] = str(row[1]).strip()
        result["config"] = config
    else:
        result["config"] = {
            "title": "비즈니스 성장 전략",
            "subtitle": "_디지털 전환을 통한 시장 확대 방안",
            "date": "2025.10.24",
            "company": "Business proposal",
            "year": "2025",
        }

    # ── slides (order/type manifest) ─────────────────────────────────────────
    if "slides" in sheets:
        result["slides"] = sheet_to_records(sheets["slides"])
    else:
        # 기본 슬라이드 순서 (나머지 시트에서 자동 추론)
        result["slides"] = None  # layout planner가 결정

    # ── 각 슬라이드 타입별 데이터 ────────────────────────────────────────────
    for sheet_type in ["stat", "two_col", "steps", "metrics", "chart", "closing"]:
        if sheet_type in sheets:
            result[sheet_type] = sheet_to_records(sheets[sheet_type])

    # table 시트는 section_num으로 그룹핑 (한 섹션 = 한 슬라이드)
    if "table" in sheets:
        result["table"] = _parse_table_sheet(sheets["table"])

    return result


def auto_build_slide_plan(data: dict) -> list:
    """slides 시트가 없을 때 데이터 시트에서 슬라이드 계획 자동 생성"""
    plan = []
    order = 1

    # Cover는 항상 첫 슬라이드
    plan.append({"order": order, "type": "cover", "section_num": "", "title": ""})
    order += 1

    # 각 데이터 시트 순서대로 슬라이드 생성
    type_sequence = ["stat", "two_col", "steps", "metrics", "chart", "table"]
    section_counter = 1

    for stype in type_sequence:
        if stype not in data:
            continue
        for rec in data[stype]:
            sec_num = str(rec.get("section_num", f"{section_counter:02d}")).zfill(2)
            title = str(rec.get("title", rec.get("section_title", "")))
            plan.append({
                "order": order,
                "type": stype,
                "section_num": sec_num,
                "title": title,
                "data": rec,
            })
            order += 1
            section_counter += 1

    # TOC는 cover 다음에 삽입
    toc_items = [s for s in plan if s["type"] not in ("cover",)]
    plan.insert(1, {
        "order": 1.5,
        "type": "toc",
        "section_num": "",
        "title": "목차",
        "items": [{"num": f"{i+1:02d}", "text": s["title"]} for i, s in enumerate(toc_items)],
    })

    # Closing은 항상 마지막
    if "closing" in data and data["closing"]:
        rec = data["closing"][0]
    else:
        rec = {"body": "우리의 비전을 실현하기 위한 구체적인 실행 계획을 제시했습니다."}
    plan.append({
        "order": order,
        "type": "closing",
        "section_num": f"{section_counter:02d}",
        "title": "결론",
        "data": rec,
    })

    # order 재정렬
    plan.sort(key=lambda x: float(x["order"]))
    for i, s in enumerate(plan):
        s["order"] = i + 1

    return plan


def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_excel.py <excel_path> [output_json_path]")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "content.json"

    print(f"[analyze_excel] Parsing: {excel_path}")
    data = parse_excel(excel_path)

    # slides 시트 없으면 자동 계획 수립
    if not data.get("slides"):
        data["slide_plan"] = auto_build_slide_plan(data)
        print(f"[analyze_excel] Auto-generated {len(data['slide_plan'])} slides")
    else:
        data["slide_plan"] = None  # layout planner가 처리

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"[analyze_excel] Output: {output_path}")
    return data


if __name__ == "__main__":
    main()
